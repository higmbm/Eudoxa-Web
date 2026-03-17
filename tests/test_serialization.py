import logging
import unittest
import eudoxa
from eudoxa import (
    Aspect, Consequence, VDiff, EudoxaManager,
    str_to_type
)

logging.getLogger("eudoxa").setLevel(logging.WARNING)


# -------------------------------------------------------------
# Helpers: build managers used across multiple tests
# -------------------------------------------------------------

def build_konsert_manager():
    """Build the 'konsert' manager (mirrors konsert.xlsx)."""
    mgr = EudoxaManager()
    mgr.add_aspect("Taxikostnad", "int", "Extrakostnad för taxi")
    mgr.add_aspect("Betyg", "str", "Betyg på tentamen")
    mgr.add_aspect("Konsert", "str", "Konsertupplevelse")

    mgr.add_aspect_level("Taxikostnad", "0",   "0 kr")
    mgr.add_aspect_level("Taxikostnad", "600", "600 kr")
    mgr.add_aspect_level("Betyg", "VG", None)
    mgr.add_aspect_level("Betyg", "G",  None)
    mgr.add_aspect_level("Betyg", "IG", None)
    mgr.add_aspect_level("Konsert", "K-", "Ingen konsert")
    mgr.add_aspect_level("Konsert", "K+", "Konsert")

    return mgr


def build_demo_manager():
    """Build a simple manager with two aspects and two consequences."""
    mgr = EudoxaManager()
    mgr.add_aspect("Kvalitet", "int", "Hur bra något är")
    mgr.add_aspect("Pris", "float", "Kostnadsnivå")

    mgr.add_aspect_level("Kvalitet", "1", "Låg kvalitet")
    mgr.add_aspect_level("Kvalitet", "2", "Medel kvalitet")
    mgr.add_aspect_level("Pris", "10",  "Billigt")
    mgr.add_aspect_level("Pris", "100", "Dyrt")

    mgr.add_consequence("C1", {"Kvalitet": "1", "Pris": "10"})
    mgr.add_consequence("C2", {"Kvalitet": "2", "Pris": "100"})

    return mgr


# -------------------------------------------------------------
# TEST: Serialization
# -------------------------------------------------------------

class TestSerialization(unittest.TestCase):

    def test_aspect_serialization(self):
        asp = Aspect("Kvalitet", int, "Beskrivning här")
        asp.levels = {"1": "Låg", "2": "Hög"}
        asp.vdiffs = [
            VDiff("Kvalitet", None, None),
            VDiff("Kvalitet", "1", "2"),
            VDiff("Kvalitet", "2", "1")
        ]

        d = asp.to_dict()
        asp2 = Aspect.from_dict(d)

        self.assertEqual(asp2.name, "Kvalitet")
        self.assertEqual(asp2.data_type, int)
        self.assertEqual(asp2.description, "Beskrivning här")
        self.assertEqual(asp2.levels, {"1": "Låg", "2": "Hög"})
        self.assertEqual(len(asp2.vdiffs), 3)
        self.assertEqual(asp2.vdiffs[1].from_level, "1")
        self.assertEqual(asp2.vdiffs[1].to_level, "2")

    def test_consequence_serialization(self):
        c = Consequence({"A": "1", "B": "x"})
        d = c.to_dict()
        c2 = Consequence.from_dict(d)

        self.assertEqual(c2.aspect_levels, {"A": "1", "B": "x"})

    def test_vdiff_cm_serialization_in_manager(self):
        mgr = build_demo_manager()
        d = mgr.to_dict()
        mgr2 = EudoxaManager.from_dict(d)

        self.assertEqual(
            set(mgr.vdiff_comparison_matrix.keys()),
            set(mgr2.vdiff_comparison_matrix.keys())
        )
        # Spot-check the first key
        for key in mgr.vdiff_comparison_matrix.keys():
            self.assertIn(key, mgr2.vdiff_comparison_matrix)
            self.assertEqual(
                set(mgr.vdiff_comparison_matrix[key].keys()),
                set(mgr2.vdiff_comparison_matrix[key].keys())
            )
            break

    def test_consequence_space_serialization(self):
        mgr = build_demo_manager()
        d = mgr.to_dict()
        mgr2 = EudoxaManager.from_dict(d)

        self.assertEqual(len(mgr.consequence_space), len(mgr2.consequence_space))
        for c1, c2 in zip(mgr.consequence_space, mgr2.consequence_space):
            self.assertEqual(c1.aspect_levels, c2.aspect_levels)

    def test_manager_full_roundtrip(self):
        mgr = build_demo_manager()
        d = mgr.to_dict()
        mgr2 = EudoxaManager.from_dict(d)

        self.assertEqual(mgr.aspects.keys(), mgr2.aspects.keys())
        for name in mgr.aspects:
            self.assertEqual(mgr.aspects[name].levels, mgr2.aspects[name].levels)
            self.assertEqual(len(mgr.aspects[name].vdiffs), len(mgr2.aspects[name].vdiffs))

        self.assertEqual(mgr.consequences.keys(), mgr2.consequences.keys())
        self.assertEqual(len(mgr.consequence_space), len(mgr2.consequence_space))
        self.assertEqual(
            set(mgr.vdiff_comparison_matrix.keys()),
            set(mgr2.vdiff_comparison_matrix.keys())
        )


# -------------------------------------------------------------
# TEST: Type validation in Aspect.add_level  (Step 1)
# -------------------------------------------------------------

class TestTypeValidation(unittest.TestCase):

    # --- int aspect ---

    def test_int_level_valid(self):
        """Integer strings are accepted on an int aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Cost", "int")
        mgr.add_aspect_level("Cost", "0",   "Zero")
        mgr.add_aspect_level("Cost", "600", "Six hundred")
        self.assertEqual(list(mgr.aspects["Cost"].levels.keys()), ["0", "600"])

    def test_int_level_rejects_non_integer(self):
        """A non-integer string must raise ValueError on an int aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Cost", "int")
        with self.assertRaises(ValueError):
            mgr.add_aspect_level("Cost", "abc", None)

    def test_int_level_rejects_float_string(self):
        """A float string (e.g. '1.5') must raise ValueError on an int aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Cost", "int")
        with self.assertRaises(ValueError):
            mgr.add_aspect_level("Cost", "1.5", None)

    # --- float aspect ---

    def test_float_level_valid(self):
        """Integer and float strings are both accepted on a float aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Price", "float")
        mgr.add_aspect_level("Price", "10",  "Ten")
        mgr.add_aspect_level("Price", "1.5", "One point five")
        self.assertEqual(list(mgr.aspects["Price"].levels.keys()), ["10", "1.5"])

    def test_float_level_rejects_non_numeric(self):
        """A non-numeric string must raise ValueError on a float aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Price", "float")
        with self.assertRaises(ValueError):
            mgr.add_aspect_level("Price", "expensive", None)

    # --- str aspect ---

    def test_str_level_accepts_anything(self):
        """Any string is valid on a str aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Grade", "str")
        mgr.add_aspect_level("Grade", "VG", None)
        mgr.add_aspect_level("Grade", "G",  None)
        mgr.add_aspect_level("Grade", "IG", None)
        self.assertEqual(list(mgr.aspects["Grade"].levels.keys()), ["VG", "G", "IG"])

    def test_str_level_accepts_numeric_string(self):
        """Numeric strings are also valid on a str aspect."""
        mgr = EudoxaManager()
        mgr.add_aspect("Code", "str")
        mgr.add_aspect_level("Code", "42", None)
        self.assertIn("42", mgr.aspects["Code"].levels)

    # --- konsert fixture ---

    def test_konsert_manager_builds_without_error(self):
        """The full konsert example must build without any type errors."""
        mgr = build_konsert_manager()
        self.assertEqual(set(mgr.aspects.keys()), {"Taxikostnad", "Betyg", "Konsert"})
        self.assertEqual(list(mgr.aspects["Taxikostnad"].levels.keys()), ["0", "600"])
        self.assertEqual(list(mgr.aspects["Betyg"].levels.keys()),       ["VG", "G", "IG"])
        self.assertEqual(list(mgr.aspects["Konsert"].levels.keys()),     ["K-", "K+"])

    def test_konsert_int_aspect_rejects_bad_level(self):
        """Adding a non-integer level to Taxikostnad (int) must raise ValueError."""
        mgr = build_konsert_manager()
        with self.assertRaises(ValueError):
            mgr.add_aspect_level("Taxikostnad", "gratis", None)


if __name__ == "__main__":
    unittest.main()

# -------------------------------------------------------------
# TEST: Aspect level relations round-trip  (Step 2)
# -------------------------------------------------------------

class TestRelationsSerialization(unittest.TestCase):

    def test_relations_survive_serialization(self):
        """Relations set before serialization must be identical after round-trip."""
        mgr = EudoxaManager()
        mgr.add_aspect("Betyg", "str")
        mgr.add_aspect_level("Betyg", "VG", None)
        mgr.add_aspect_level("Betyg", "G",  None)
        mgr.add_aspect_level("Betyg", "IG", None)
        mgr.set_aspect_level_relation("Betyg", "VG", "G",  eudoxa.BT)
        mgr.set_aspect_level_relation("Betyg", "G",  "IG", eudoxa.BT)

        mgr2 = EudoxaManager.from_dict(mgr.to_dict())

        self.assertEqual(mgr2.get_aspect_level_relation("Betyg", "VG", "G"),  eudoxa.BT)
        self.assertEqual(mgr2.get_aspect_level_relation("Betyg", "G",  "IG"), eudoxa.BT)
        self.assertEqual(mgr2.get_aspect_level_relation("Betyg", "VG", "IG"), eudoxa.UNDEFINED)

    def test_konsert_relations_survive_serialization(self):
        """Relations from the konsert fixture must survive a round-trip."""
        mgr = build_konsert_manager()
        mgr.set_aspect_level_relation("Taxikostnad", "0",   "600", eudoxa.BT)
        mgr.set_aspect_level_relation("Betyg",       "VG",  "G",   eudoxa.BT)
        mgr.set_aspect_level_relation("Betyg",       "G",   "IG",  eudoxa.BT)
        mgr.set_aspect_level_relation("Konsert",     "K+",  "K-",  eudoxa.BT)

        mgr2 = EudoxaManager.from_dict(mgr.to_dict())

        self.assertEqual(mgr2.get_aspect_level_relation("Taxikostnad", "0",  "600"), eudoxa.BT)
        self.assertEqual(mgr2.get_aspect_level_relation("Taxikostnad", "600", "0"),  eudoxa.WT)
        self.assertEqual(mgr2.get_aspect_level_relation("Betyg", "VG", "G"),         eudoxa.BT)
        self.assertEqual(mgr2.get_aspect_level_relation("Konsert", "K+", "K-"),      eudoxa.BT)


# -------------------------------------------------------------
# TEST: Import from worksheet  (Step 2 - foundation for Step 3)
# -------------------------------------------------------------

class TestImportFromWorksheet(unittest.TestCase):

    def _make_aspect_worksheet(self, wb, name, data_type_str, description, levels):
        """Helper: write an aspect worksheet in the format EudoxaManager expects."""
        import openpyxl
        ws = wb.create_sheet(title=eudoxa.ASP + name)
        ws["A1"] = name
        ws["B1"] = data_type_str
        ws["A2"] = description
        for row_idx, (level, desc) in enumerate(levels, start=3):
            ws.cell(row=row_idx, column=1).value = level
            ws.cell(row=row_idx, column=2).value = desc
        return ws

    def test_import_aspect_levels(self):
        """import_aspect_from_worksheet must recreate levels correctly."""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._make_aspect_worksheet(
            wb, "Betyg", "str", "Betyg pa tentamen",
            [("VG", None), ("G", None), ("IG", None)]
        )

        mgr = EudoxaManager()
        mgr.import_aspect_from_worksheet(wb[eudoxa.ASP + "Betyg"])

        self.assertIn("Betyg", mgr.aspects)
        self.assertEqual(list(mgr.aspects["Betyg"].levels.keys()), ["VG", "G", "IG"])
        self.assertEqual(mgr.aspects["Betyg"].description, "Betyg pa tentamen")

    def test_import_int_aspect_rejects_bad_level(self):
        """Importing an int aspect with a non-integer level must raise ValueError."""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._make_aspect_worksheet(
            wb, "Kostnad", "int", None,
            [("100", "Hundra"), ("gratis", "Ogiltigt")]
        )

        mgr = EudoxaManager()
        with self.assertRaises(ValueError):
            mgr.import_aspect_from_worksheet(wb[eudoxa.ASP + "Kostnad"])

    def test_import_consequences_from_worksheet(self):
        """import_consequences_from_worksheet must recreate consequences correctly."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = eudoxa.CONS
        ws["A1"] = None;  ws["B1"] = "Taxikostnad"; ws["C1"] = "Betyg"
        ws["A2"] = None;  ws["B2"] = "int";          ws["C2"] = "str"
        ws["A3"] = "C11"; ws["B3"] = 0;   ws["C3"] = "VG"
        ws["A4"] = "C12"; ws["B4"] = 0;   ws["C4"] = "G"
        ws["A5"] = "C21"; ws["B5"] = 600; ws["C5"] = "VG"

        mgr = EudoxaManager()
        mgr.import_consequences_from_worksheet(ws)

        self.assertIn("C11", mgr.consequences)
        self.assertIn("C12", mgr.consequences)
        self.assertIn("C21", mgr.consequences)
        self.assertEqual(mgr.consequences["C11"]["Taxikostnad"], "0")
        self.assertEqual(mgr.consequences["C21"]["Betyg"], "VG")


# -------------------------------------------------------------
# TEST: Consequence space derivability  (Step 2 - foundation for Step 4)
# -------------------------------------------------------------

class TestConsequenceSpaceDerivability(unittest.TestCase):

    def _derive_consequence_space(self, mgr):
        """
        Recompute the consequence space from scratch using only
        aspects and their levels - no reference to mgr.consequence_space.
        This mirrors what Step 4 will do when dropping it from the session.
        """
        from itertools import product
        aspects = list(mgr.aspects.values())
        if not aspects:
            return [Consequence()]
        level_lists = [list(a.levels.keys()) for a in aspects]
        space = []
        for combo in product(*level_lists):
            c = Consequence({a.name: level for a, level in zip(aspects, combo)})
            space.append(c)
        return space

    def test_consequence_space_matches_derived(self):
        """The stored consequence_space must equal the space derived from aspects+levels."""
        mgr = build_demo_manager()
        derived = self._derive_consequence_space(mgr)

        stored_sets  = [frozenset(c.aspect_levels.items()) for c in mgr.consequence_space]
        derived_sets = [frozenset(c.aspect_levels.items()) for c in derived]

        self.assertEqual(sorted(stored_sets,  key=lambda fs: str(sorted(fs))),
                         sorted(derived_sets, key=lambda fs: str(sorted(fs))))

    def test_consequence_space_matches_derived_konsert(self):
        """Consequence space for the konsert fixture must match the derived space."""
        mgr = build_konsert_manager()
        derived = self._derive_consequence_space(mgr)

        stored_sets  = [frozenset(c.aspect_levels.items()) for c in mgr.consequence_space]
        derived_sets = [frozenset(c.aspect_levels.items()) for c in derived]

        self.assertEqual(sorted(stored_sets,  key=lambda fs: str(sorted(fs))),
                         sorted(derived_sets, key=lambda fs: str(sorted(fs))))

    def test_consequence_space_size(self):
        """Consequence space size must equal the product of the number of levels per aspect."""
        mgr = build_konsert_manager()
        expected_size = 2 * 3 * 2  # Taxikostnad x Betyg x Konsert
        self.assertEqual(len(mgr.consequence_space), expected_size)

    def test_empty_manager_consequence_space(self):
        """A manager with no aspects must have a consequence space of size 1 (empty consequence)."""
        mgr = EudoxaManager()
        self.assertEqual(len(mgr.consequence_space), 1)
        self.assertEqual(mgr.consequence_space[0].aspect_levels, {})

# -------------------------------------------------------------
# TEST: try_set_aspect_level_relation  (Step 5)
# -------------------------------------------------------------

class TestTrySetAspectLevelRelation(unittest.TestCase):

    def _grade_manager(self):
        """Three-level aspect: VG > G > IG (by convention)."""
        mgr = EudoxaManager()
        mgr.add_aspect("Grade", "str")
        mgr.add_aspect_level("Grade", "VG", None)
        mgr.add_aspect_level("Grade", "G",  None)
        mgr.add_aspect_level("Grade", "IG", None)
        return mgr

    # ── i. Clean case ─────────────────────────────────────────

    def test_clean_addition_succeeds(self):
        """A non-conflicting relation is committed; no collisions returned."""
        mgr = self._grade_manager()
        adds, colls, inferred_adds = mgr.try_set_aspect_level_relation(
            "Grade", "VG", "G", eudoxa.BT
        )
        self.assertEqual(colls, [], "Expected no collisions")
        self.assertGreater(len(adds), 0, "Expected at least one direct add")
        self.assertEqual(
            mgr.get_aspect_level_relation("Grade", "VG", "G"),
            eudoxa.BT,
            "Relation should be committed to matrix"
        )

    def test_clean_addition_reports_inferred(self):
        """After VG>G and G>IG are set, inferred_adds should include VG>IG."""
        mgr = self._grade_manager()
        mgr.try_set_aspect_level_relation("Grade", "VG", "G",  eudoxa.BT)
        adds, colls, inferred_adds = mgr.try_set_aspect_level_relation(
            "Grade", "G",  "IG", eudoxa.BT
        )
        self.assertEqual(colls, [])
        self.assertGreater(len(inferred_adds), 0,
            "Expected inferred additions after transitive chain")

    def test_clean_addition_does_not_modify_matrix_beyond_explicit(self):
        """Only the explicit addition is written to the matrix; inferred
        relations are NOT stored (they live only in the closure)."""
        mgr = self._grade_manager()
        mgr.try_set_aspect_level_relation("Grade", "VG", "G",  eudoxa.BT)
        mgr.try_set_aspect_level_relation("Grade", "G",  "IG", eudoxa.BT)
        # VG > IG is inferred but must NOT be explicitly in the matrix
        # (it is only derivable via the closure)
        vd_vg_ig = eudoxa.VDiff("Grade", "VG", "IG")
        vd_zero  = eudoxa.VDiff("Grade", None, None)
        raw = eudoxa.get_vdiff_relation(
            mgr.vdiff_comparison_matrix, vd_vg_ig, vd_zero
        )
        self.assertEqual(raw, eudoxa.UNDEFINED,
            "Inferred relation VG>IG should not be written to the matrix")

    # ── ii. Immediate collision ───────────────────────────────

    def test_immediate_collision_rejected(self):
        """Asserting a relation that directly contradicts the current closure
        is rejected; no changes are made to the matrix."""
        mgr = self._grade_manager()
        mgr.try_set_aspect_level_relation("Grade", "VG", "G", eudoxa.BT)

        # G > VG directly contradicts VG > G
        adds, colls, inferred_adds = mgr.try_set_aspect_level_relation(
            "Grade", "G", "VG", eudoxa.BT
        )
        self.assertGreater(len(colls), 0, "Expected collision to be reported")
        self.assertEqual(adds, [], "No adds should be returned on collision")

    def test_immediate_collision_leaves_matrix_unchanged(self):
        """The matrix must be identical before and after a rejected addition."""
        mgr = self._grade_manager()
        mgr.try_set_aspect_level_relation("Grade", "VG", "G", eudoxa.BT)

        import copy
        matrix_before = copy.deepcopy(mgr.vdiff_comparison_matrix)

        mgr.try_set_aspect_level_relation("Grade", "G", "VG", eudoxa.BT)

        self.assertEqual(
            mgr.vdiff_comparison_matrix, matrix_before,
            "Matrix must be unchanged after a rejected addition"
        )

    def test_transitive_collision_rejected(self):
        """A relation that creates a cycle via transitivity (VG>G>IG, then IG>VG)
        is caught because the current closure already contains VG>IG.
        The collision is therefore detected at the immediate-check stage."""
        mgr = self._grade_manager()
        mgr.try_set_aspect_level_relation("Grade", "VG", "G",  eudoxa.BT)
        mgr.try_set_aspect_level_relation("Grade", "G",  "IG", eudoxa.BT)

        adds, colls, inferred_adds = mgr.try_set_aspect_level_relation(
            "Grade", "IG", "VG", eudoxa.BT
        )
        self.assertGreater(len(colls), 0,
            "Cyclic relation should be rejected")
        self.assertEqual(
            mgr.get_aspect_level_relation("Grade", "VG", "G"),
            eudoxa.BT,
            "Previously committed relation must be intact after rejection"
        )

    # ── iii. Inferred collision via closure recomputation ─────
    #
    # Within a single aspect's level relations, any cycle is caught at the
    # immediate-check stage (step 2) rather than the recomputed-closure stage
    # (step 5), because the current closure already contains all transitively
    # inferred relations before the addition is attempted.
    #
    # The step-5 path is reachable via cross-aspect vdiff relations (set_rel),
    # which is not yet exposed through try_set_aspect_level_relation.
    # The closure() method itself is tested below to confirm it detects
    # inferred collisions when cross-aspect relations form a contradiction.

    def test_closure_detects_inferred_collision(self):
        """The closure() method must report a collision when transitivity
        across three aspects produces a contradiction not detectable from
        any single pair of explicitly set vdiff relations.

        Setup: X >= Y and Y >= Z are set explicitly.
        Transitivity infers X >= Z.
        Then Z > X (strictly) is set explicitly — each individual
        assertion is accepted, but the combination is contradictory.
        """
        mgr = EudoxaManager()
        for name in ["X", "Y", "Z"]:
            mgr.add_aspect(name, "str")
            mgr.add_aspect_level(name, f"{name.lower()}1", None)
            mgr.add_aspect_level(name, f"{name.lower()}2", None)

        vd_x  = eudoxa.VDiff("X", "x1", "x2")
        vd_y  = eudoxa.VDiff("Y", "y1", "y2")
        vd_z  = eudoxa.VDiff("Z", "z1", "z2")
        vdcm  = mgr.vdiff_comparison_matrix

        # X >= Y, Y >= Z — individually consistent
        eudoxa.set_vdiff_relation(vdcm, vd_x, vd_y, eudoxa.TRUE)
        eudoxa.set_vdiff_relation(vdcm, vd_y, vd_z, eudoxa.TRUE)
        # Z strictly > X — accepted by set_vdiff_relation since the
        # transitive contradiction (X>=Y>=Z => X>=Z) is not checked there
        eudoxa.set_vdiff_relation(vdcm, vd_z, vd_x, eudoxa.TRUE)
        eudoxa.set_vdiff_relation(vdcm, vd_x, vd_z, eudoxa.FALSE)

        _, _, closure_colls = mgr.closure()
        self.assertGreater(len(closure_colls), 0,
            "Closure must detect the transitive contradiction across three aspects")